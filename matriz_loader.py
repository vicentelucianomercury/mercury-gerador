"""
matriz_loader.py — Mercury · Leitura da planilha da Matriz
===========================================================
Lê as abas "Matriz Editavel" (layout em blocos, posicional) e
"Lista de Fundos Aprovados" (por nome de cabeçalho) de um Excel
enviado pelo usuário, valida e devolve a configuração no MESMO
formato do fundos_config_v2 (FUNDOS / PERFIS_TATICA / PERFIS_BANDAS /
PERFIS_METAS).

Regras (decisão do comitê, jul/2026):
- Vocabulário de classes FECHADO: as 10 classes canônicas do código.
  A planilha muda números e fundos; classe nova exige desenvolvedor.
- Erros travam (o app não deixa avançar); avisos não.
- Fundo sem série no arquivo de cotas é AVISO, não erro
  (fundo novo pode ainda não ter histórico).
- Casamento fundo↔série por nome normalizado com prefixo/continência
  (os nomes vêm truncados em ambos os lados).

Este módulo é Python puro (openpyxl) e não importa streamlit,
para poder ser testado fora do app.
"""
from __future__ import annotations

import io
import re
import unicodedata
from dataclasses import dataclass, field

import openpyxl

# ─── Classes canônicas (vocabulário fechado — igual a fundos_config_v2) ──────
CLASSES_CANONICAS = [
    "Caixa",
    "Crédito Corporativo",
    "Crédito Estruturado",
    "Renda Fixa Inflação/Pré",
    "Multimercado",
    "Offshore USD",
    "Offshore BRL",
    "Renda Variável",
    "Imobiliários",
    "Private Equity/VC",
]

ABA_MATRIZ = "Matriz Editavel"
ABA_FUNDOS = "Lista de Fundos Aprovados"

PERFIS_ESPERADOS = ["Conservador", "Moderado", "Balanceado",
                    "Crescimento", "Sofisticado"]

# Offsets dos dois blocos por faixa: (coluna do rótulo, 1ª coluna de dados)
# Bloco esquerdo: rótulos na col A (1), dados em B..F (2..6)
# Bloco direito:  rótulos na col H (8), dados em I..M (9..13)
_BLOCOS_COL = [(1, 2), (8, 9)]


# ─── Normalização ─────────────────────────────────────────────────────────────
def _norm(s) -> str:
    """MAIÚSCULAS, sem acento, espaços colapsados."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", s).strip().upper()


# Aliases → classe canônica. Casados por prefixo bidirecional
# (rótulos truncados: "CRED CORPORATI", "RENDA VARIAVEL L", "PRIVATE EQUITY/V").
# Ordenados do mais específico para o mais genérico — "RENDA FIXA INFLACAO"
# precisa vir antes de qualquer coisa que comece com "RENDA".
_ALIASES: list[tuple[str, str]] = [
    ("RENDA FIXA INFLACAO/PRE", "Renda Fixa Inflação/Pré"),
    ("RENDA FIXA INFLACAO",     "Renda Fixa Inflação/Pré"),
    ("PRIVATE EQUITY/VC",       "Private Equity/VC"),
    ("PRIVATE EQUITY",          "Private Equity/VC"),
    ("CREDITO CORPORATIVO",     "Crédito Corporativo"),
    ("CRED CORPORATIVO",        "Crédito Corporativo"),
    ("CREDITO ESTRUTURADO",     "Crédito Estruturado"),
    ("RENDA VARIAVEL",          "Renda Variável"),
    ("ACOES",                   "Renda Variável"),
    ("MULTIMERCADOS",           "Multimercado"),
    ("MULTIMERCADO",            "Multimercado"),
    ("OFFSHORE USD",            "Offshore USD"),
    ("OFFSHORE BRL",            "Offshore BRL"),
    ("IMOBILIARIOS",            "Imobiliários"),
    ("IMOBILIARIO",             "Imobiliários"),
    ("CAIXA",                   "Caixa"),
]

_MIN_PREFIXO = 5  # evita casar por acidente com rótulos muito curtos


def classe_canonica(rotulo) -> str | None:
    """Mapeia um rótulo (possivelmente truncado/variação) para a classe
    canônica. Retorna None se não reconhecer."""
    n = _norm(rotulo)
    if not n:
        return None
    for alias, canon in _ALIASES:
        if len(n) >= _MIN_PREFIXO and (n.startswith(alias) or alias.startswith(n)):
            return canon
    return None


def _num(v) -> float | None:
    """Converte célula em fração. Aceita 0.185, '18,5%', '18,5'. None → None."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("%", "").replace(",", ".")
    try:
        x = float(s)
    except ValueError:
        return None
    return x / 100 if x > 1.5 else x  # '18,5' veio como número inteiro-percentual


# ─── Resultado ────────────────────────────────────────────────────────────────
@dataclass
class MatrizConfig:
    fundos: list[dict] = field(default_factory=list)      # {nome_quantum,cnpj,classe[,subclasse]}
    tatica: dict = field(default_factory=dict)            # perfil -> {classe: frac}
    bandas: dict = field(default_factory=dict)            # perfil -> {classe: (min,max)}
    metas:  dict = field(default_factory=dict)            # perfil -> {objetivo,vol}
    erros:  list[str] = field(default_factory=list)       # travam
    avisos: list[str] = field(default_factory=list)       # não travam
    origem: str = ""                                      # nome do arquivo

    @property
    def valida(self) -> bool:
        return not self.erros


# ─── Leitura da aba "Lista de Fundos Aprovados" ───────────────────────────────
def _ler_lista_fundos(wb, cfg: MatrizConfig) -> None:
    ws = wb[ABA_FUNDOS]

    # Cabeçalhos por NOME (tolerando acento/caixa), não por posição
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = _norm(ws.cell(row=1, column=c).value)
        if h:
            headers[h] = c

    def col(*nomes):
        for nm in nomes:
            if nm in headers:
                return headers[nm]
        return None

    c_fundo = col("FUNDO", "NOME", "NOME DO FUNDO")
    c_cnpj  = col("CNPJ")
    c_cls   = col("CLASSE DE ATIVO", "CLASSE", "CLASSE DE ATIVOS")
    c_sub   = col("SUBCLASSE", "SUB CLASSE", "SUB-CLASSE")   # opcional

    faltando = [nm for nm, c in
                [("FUNDO", c_fundo), ("CNPJ", c_cnpj), ("CLASSE DE ATIVO", c_cls)]
                if c is None]
    if faltando:
        cfg.erros.append(
            f"Na aba \u201c{ABA_FUNDOS}\u201d não encontrei as colunas: "
            f"{', '.join(faltando)}. Confira a linha 1 (cabeçalho).")
        return

    sem_cnpj, cls_invalida = [], []
    for r in range(2, ws.max_row + 1):
        nome = ws.cell(row=r, column=c_fundo).value
        if nome is None or not str(nome).strip():
            continue  # linha vazia
        nome = str(nome).strip()
        cnpj = ws.cell(row=r, column=c_cnpj).value
        cnpj = str(cnpj).strip() if cnpj not in (None, "") else ""
        rot_cls = ws.cell(row=r, column=c_cls).value
        canon = classe_canonica(rot_cls)

        if not cnpj:
            sem_cnpj.append(nome)
        if canon is None:
            cls_invalida.append(f"\u201c{rot_cls}\u201d (fundo {nome[:40]}…)"
                                if rot_cls else f"vazia (fundo {nome[:40]}…)")
            continue

        item = {"nome_quantum": nome, "cnpj": cnpj, "classe": canon}
        if c_sub is not None:
            sub = ws.cell(row=r, column=c_sub).value
            if sub not in (None, ""):
                item["subclasse"] = str(sub).strip()
        cfg.fundos.append(item)

    if sem_cnpj:
        cfg.erros.append(
            "Fundos sem CNPJ na aba \u201cLista de Fundos Aprovados\u201d: "
            + "; ".join(n[:50] for n in sem_cnpj)
            + ". Preencha o CNPJ e envie novamente.")
    if cls_invalida:
        cfg.erros.append(
            "Classes não reconhecidas na Lista de Fundos: "
            + "; ".join(cls_invalida)
            + ". Classes válidas: " + ", ".join(CLASSES_CANONICAS)
            + ". Corrija a planilha ou peça ajuste no sistema "
              "(criar classe nova exige desenvolvedor).")
    if not cfg.fundos and not cfg.erros:
        cfg.erros.append("A aba \u201cLista de Fundos Aprovados\u201d está vazia.")


# ─── Leitura da aba "Matriz Editavel" (blocos) ────────────────────────────────
def _ler_matriz(wb, cfg: MatrizConfig) -> None:
    ws = wb[ABA_MATRIZ]
    max_r = ws.max_row

    def val(r, c):
        return ws.cell(row=r, column=c).value

    # Âncora: linha-cabeçalho do bloco ("Min | Base | Max | Tática | Var").
    # As faixas de linha da descrição (1-15 / 17-31 / 33+) são apenas o layout
    # típico; o Crescimento, p.ex., termina na linha 32 por ter PE/VC. Por isso
    # os blocos são localizados pela âncora, não por linha fixa.
    blocos = []  # (linha_header, col_rotulo, col_dados0, nome_perfil)
    for r in range(2, max_r + 1):
        for col_rot, col_d in _BLOCOS_COL:
            if (_norm(val(r, col_d)) == "MIN"
                    and _norm(val(r, col_d + 1)) == "BASE"
                    and _norm(val(r, col_d + 3)).startswith("TATIC")):
                perfil = val(r - 1, col_d)
                perfil = str(perfil).strip() if perfil else f"(sem nome, linha {r-1})"
                blocos.append((r, col_rot, col_d, perfil))

    if not blocos:
        cfg.erros.append(
            f"Não encontrei nenhum bloco de perfil na aba \u201c{ABA_MATRIZ}\u201d "
            "(procurei pelo cabeçalho \u201cMin | Base | Max | Tática | Var\u201d). "
            "Confira se a aba mantém o layout em blocos.")
        return

    for r_head, c_rot, c_d, perfil in blocos:
        tatica, bandas = {}, {}
        objetivo, vol_meta = None, None
        rotulos_invalidos = []

        r = r_head + 1
        while r <= max_r:
            rot = val(r, c_rot)
            n = _norm(rot)
            if n == "TOTAL":
                # Objetivo e Volatilidade nas linhas seguintes do bloco
                for rr in range(r + 1, min(r + 4, max_r + 1)):
                    n2 = _norm(val(rr, c_rot))
                    if n2.startswith("OBJETIVO"):
                        v = val(rr, c_d)
                        objetivo = str(v).strip() if v not in (None, "") else None
                    elif n2.startswith("VOLATILIDADE"):
                        vol_meta = _num(val(rr, c_d))
                break
            if not n:
                r += 1
                continue
            if n == "RENDA FIXA":          # linha-grupo: só Base, não é alocável
                r += 1
                continue
            canon = classe_canonica(rot)
            if canon is None:
                rotulos_invalidos.append(str(rot).strip())
                r += 1
                continue
            mn = _num(val(r, c_d))          # Min
            mx = _num(val(r, c_d + 2))      # Max
            tt = _num(val(r, c_d + 3))      # Tática (vazio = 0)
            bandas[canon] = (mn or 0.0, mx or 0.0)
            tatica[canon] = tt or 0.0
            r += 1
        else:
            cfg.avisos.append(
                f"Perfil \u201c{perfil}\u201d: não encontrei a linha \u201cTotal\u201d "
                "ao fim do bloco — li até o fim da aba.")

        if rotulos_invalidos:
            cfg.erros.append(
                f"Perfil \u201c{perfil}\u201d: classe(s) não reconhecida(s): "
                + "; ".join(f"\u201c{x}\u201d" for x in rotulos_invalidos)
                + ". Classes válidas: " + ", ".join(CLASSES_CANONICAS) + ".")

        # Classes ausentes no bloco entram zeradas (perfis sem PE/VC, p.ex.)
        for c in CLASSES_CANONICAS:
            tatica.setdefault(c, 0.0)
            bandas.setdefault(c, (0.0, 0.0))

        soma = sum(tatica.values())
        if abs(soma - 1.0) > 0.001:  # tolerância 0,1 p.p.
            cfg.erros.append(
                f"A alocação tática do perfil \u201c{perfil}\u201d soma "
                f"{soma*100:.2f}%".replace(".", ",")
                + " — deveria somar 100% (tolerância de 0,1 p.p.). "
                  "Ajuste a coluna Tática e envie novamente.")

        cfg.tatica[perfil] = tatica
        cfg.bandas[perfil] = bandas
        cfg.metas[perfil] = {"objetivo": objetivo, "vol": vol_meta}

    achados = list(cfg.tatica.keys())
    faltam = [p for p in PERFIS_ESPERADOS if p not in achados]
    extras = [p for p in achados if p not in PERFIS_ESPERADOS]
    if faltam:
        cfg.erros.append(
            "Perfis não encontrados na Matriz Editavel: " + ", ".join(faltam)
            + f". Encontrei: {', '.join(achados) or 'nenhum'}.")
    if extras:
        cfg.avisos.append(
            "Perfis fora do padrão encontrados (serão usados mesmo assim): "
            + ", ".join(extras) + ".")


# ─── Cruzamento fundo ↔ séries de cotas ───────────────────────────────────────
def _eh_serie_valida(nome_norm: str) -> bool:
    """Filtra os disclaimers que o export do Quantum injeta como 'séries'."""
    if len(nome_norm) > 120:
        return False
    ruins = ("FONTE:", "AS INFO", "OS VAL", "AS PERFORMANCES", "FUNDOS DE",
             "PARA SEUS", "QUALQUER", "PASSADO", "E REGULAMENTO")
    return not nome_norm.startswith(ruins)


def cruzar_com_series(cfg: MatrizConfig, nomes_series) -> None:
    """Verifica quais fundos têm série de cotas. AVISO, não erro.
    Match por nome normalizado com prefixo/continência (nomes truncados)."""
    series = {}
    for s in nomes_series or []:
        n = _norm(s)
        if n and _eh_serie_valida(n):
            series[n] = str(s)

    for f in cfg.fundos:
        alvo = _norm(f["nome_quantum"])
        if alvo in series:                      # match exato normalizado
            f["serie_quantum"] = series[alvo]
            continue
        candidatos = [orig for n, orig in series.items()
                      if len(alvo) >= _MIN_PREFIXO
                      and (n.startswith(alvo) or alvo.startswith(n))]
        if len(candidatos) == 1:
            f["serie_quantum"] = candidatos[0]
            if _norm(candidatos[0]) != alvo:
                cfg.avisos.append(
                    f"Fundo \u201c{f['nome_quantum'][:45]}…\u201d casado com a série "
                    f"\u201c{candidatos[0][:45]}…\u201d por aproximação de nome.")
        elif len(candidatos) == 0:
            cfg.avisos.append(
                f"O fundo \u201c{f['nome_quantum'][:60]}\u201d não aparece no arquivo "
                "de cotas — sem série, ele não entra nos cálculos de rentabilidade "
                "(fundo novo pode ainda não ter histórico).")
        else:
            cfg.avisos.append(
                f"Não consegui casar com certeza o fundo "
                f"\u201c{f['nome_quantum'][:50]}\u201d com as séries de cotas "
                f"({len(candidatos)} candidatos). Verifique o nome na planilha.")


# ─── Entrada principal ────────────────────────────────────────────────────────
def carregar_matriz(conteudo: bytes, nome_arquivo: str = "",
                    nomes_series=None) -> MatrizConfig:
    """Lê a planilha da matriz (bytes do upload). Sempre retorna MatrizConfig;
    problemas vão em .erros (travam) e .avisos (não travam)."""
    cfg = MatrizConfig(origem=nome_arquivo or "planilha enviada")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
    except Exception as e:
        cfg.erros.append(f"Não consegui abrir o arquivo como Excel (.xlsx): {e}")
        return cfg

    for aba in (ABA_MATRIZ, ABA_FUNDOS):
        if aba not in wb.sheetnames:
            cfg.erros.append(
                f"A aba \u201c{aba}\u201d não existe na planilha. "
                f"Abas encontradas: {', '.join(wb.sheetnames)}. "
                "As demais abas (MATRIZ, Matriz New, Old Port…) são ignoradas — "
                "as duas acima precisam existir com esses nomes.")
    if cfg.erros:
        return cfg

    _ler_lista_fundos(wb, cfg)
    _ler_matriz(wb, cfg)

    if nomes_series is not None:
        cruzar_com_series(cfg, nomes_series)

    # Consistência entre abas: classe usada na Lista precisa ter linha na matriz
    # (já garantido pelo vocabulário fechado + preenchimento com zero).
    return cfg
