"""
mercury_data.py — Mercury Wealth Management
Fonte: Excel exportado via API Quantum Axis (atualização mensal/trimestral)
"""
import math, os
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl

QUANTUM_FILE = Path(os.environ.get(
    "QUANTUM_FILE", Path(__file__).parent / "data" / "quantum.xlsx"))

_cache: dict = {}

def _load_series(filepath: Path = QUANTUM_FILE) -> dict:
    key = str(filepath)
    if key in _cache: return _cache[key]
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active
    for s in wb.sheetnames:
        if "quantum" in s.lower(): ws = wb[s]; break
    series = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome, data_val, cota = row[0], row[1], row[2]
        if not (nome and data_val and isinstance(cota,(int,float))): continue
        if isinstance(data_val, datetime): data_val = data_val.date()
        elif not isinstance(data_val, date): continue
        series[str(nome).strip()].append((data_val, float(cota)))
    for nome in series: series[nome].sort(key=lambda x: x[0])
    wb.close()
    _cache[key] = dict(series)
    return _cache[key]

def _find_closest(dados, alvo):
    if not dados: return None
    c = min(dados, key=lambda x: abs((x[0]-alvo).days))
    return c if abs((c[0]-alvo).days) <= 45 else None

def _retorno_meses(dados, meses):
    if not dados: return None
    ul_data, ul_cota = dados[-1]
    m = ul_data.month - (meses % 12); y = ul_data.year - (meses // 12)
    if m <= 0: m += 12; y -= 1
    try: alvo = date(y, m, ul_data.day)
    except ValueError:
        import calendar; alvo = date(y, m, calendar.monthrange(y,m)[1])
    ref = _find_closest(dados, alvo)
    return (ul_cota/ref[1]) - 1 if ref else None

def _retorno_ytd(dados):
    if not dados: return None
    ul_data, ul_cota = dados[-1]
    alvo = date(ul_data.year-1, 12, 31)
    ref = _find_closest(dados, alvo)
    if ref is None or abs((ref[0]-alvo).days) > 10: return None
    return (ul_cota/ref[1]) - 1

def _vol_anual(dados, meses=12):
    if not dados or len(dados) < 30: return None
    ul = dados[-1][0]
    m = ul.month-(meses%12); y = ul.year-(meses//12)
    if m<=0: m+=12; y-=1
    corte = date(y,m,1)
    rec = [(d,c) for d,c in dados if d>=corte]
    if len(rec)<20: return None
    rets = [(rec[i][1]/rec[i-1][1])-1 for i in range(1,len(rec))]
    media = sum(rets)/len(rets)
    var = sum((r-media)**2 for r in rets)/len(rets)
    return math.sqrt(var*252)

def _retornos_mensais(dados):
    """Retorna dict {(ano,mes): retorno_mensal}"""
    if not dados: return {}
    by_month = {}
    for d, c in dados:
        key = (d.year, d.month)
        if key not in by_month or d > by_month[key][0]:
            by_month[key] = (d, c)
    # primeiro dia útil do mês
    first_month = {}
    for d, c in dados:
        key = (d.year, d.month)
        if key not in first_month or d < first_month[key][0]:
            first_month[key] = (d, c)
    result = {}
    months = sorted(by_month.keys())
    for i in range(1, len(months)):
        cur = months[i]; prev = months[i-1]
        c_cur  = by_month[cur][1]
        c_prev = by_month[prev][1]
        result[cur] = (c_cur/c_prev) - 1
    return result

def calcular_metricas(filepath=QUANTUM_FILE):
    series = _load_series(filepath)
    cdi = series.get("CDI",[])
    cdi_12m = _retorno_meses(cdi, 12)
    resultado = {}
    for nome, dados in series.items():
        if len(nome)>120 or nome.lower().startswith(
            ("fonte:","as info","os val","fundos de","para seus","qualquer","passado")):
            continue
        r6  = _retorno_meses(dados,6)
        r12 = _retorno_meses(dados,12)
        r24 = _retorno_meses(dados,24)
        r36 = _retorno_meses(dados,36)
        ytd = _retorno_ytd(dados)
        vol = _vol_anual(dados)
        pct = (r12/cdi_12m) if (r12 is not None and cdi_12m) else None
        resultado[nome] = {
            "ret_6m":r6,"ret_12m":r12,"ret_24m":r24,"ret_36m":r36,
            "ytd":ytd,"vol_12m":vol,"pct_cdi":pct,
            "data_inicio":dados[0][0] if dados else None,
            "data_fim":dados[-1][0] if dados else None,
            "n_pontos":len(dados),
            "retornos_mensais": _retornos_mensais(dados),
        }
    return resultado

def calcular_portfolio(aloc_fd, metricas, valor):
    """Calcula métricas agregadas da carteira ponderada pelas alocações."""
    pesos = {nq: pct/100 for nq,pct in aloc_fd.items() if pct>0}
    
    # Retorno ponderado 12m
    ret_12m = sum(pesos[nq]*metricas[nq]["ret_12m"]
                  for nq in pesos if metricas.get(nq,{}).get("ret_12m") is not None)
    ret_ytd = sum(pesos[nq]*metricas[nq]["ytd"]
                  for nq in pesos if metricas.get(nq,{}).get("ytd") is not None)
    
    # CDI 12m
    cdi_m = metricas.get("CDI",{}).get("ret_12m")
    pct_cdi = (ret_12m/cdi_m) if (cdi_m and cdi_m!=0) else None
    
    # Volatilidade ponderada (aproximação)
    vol = sum(pesos[nq]*metricas[nq]["vol_12m"]
              for nq in pesos if metricas.get(nq,{}).get("vol_12m") is not None)
    
    # Sharpe (usando CDI como risk-free)
    sharpe = ((ret_12m - (cdi_m or 0)) / vol) if vol and vol > 0 else None
    
    # Patrimônio acumulado
    patrimonio = valor * (1 + ret_12m)
    
    # Retornos mensais ponderados
    monthly = {}
    for nq, peso in pesos.items():
        m = metricas.get(nq,{}).get("retornos_mensais",{})
        for ym, r in m.items():
            monthly[ym] = monthly.get(ym,0) + peso*r
    
    # Estatísticas mensais
    cdi_monthly = metricas.get("CDI",{}).get("retornos_mensais",{})
    m_pos = sum(1 for r in monthly.values() if r>0)
    m_neg = sum(1 for r in monthly.values() if r<0)
    maior = max(monthly.values()) if monthly else None
    menor = min(monthly.values()) if monthly else None
    acima = sum(1 for ym,r in monthly.items()
                if r > cdi_monthly.get(ym,0))
    abaixo = sum(1 for ym,r in monthly.items()
                 if r <= cdi_monthly.get(ym,0))
    
    return {
        "ret_12m": ret_12m, "ret_ytd": ret_ytd,
        "vol_12m": vol, "sharpe": sharpe,
        "pct_cdi": pct_cdi, "patrimonio": patrimonio,
        "monthly": monthly, "cdi_monthly": cdi_monthly,
        "m_pos": m_pos, "m_neg": m_neg,
        "maior": maior, "menor": menor,
        "acima_cdi": acima, "abaixo_cdi": abaixo,
    }

def ultima_data_atualizacao(filepath=QUANTUM_FILE):
    series = _load_series(filepath)
    datas = [dados[-1][0] for dados in series.values() if dados]
    return max(datas) if datas else None

def cdi_acumulado(meses, filepath=QUANTUM_FILE):
    series = _load_series(filepath)
    return _retorno_meses(series.get("CDI",[]), meses)


# ─── Cálculos de portfólio (para o relatório PDF) ────────────────────────────

def _retornos_mensais(dados: list) -> dict:
    """Retorna dict {(ano,mes): retorno_mensal} a partir da série de cotas."""
    if not dados or len(dados) < 2:
        return {}
    # último ponto de cada mês
    fim_mes = {}
    for d, c in dados:
        fim_mes[(d.year, d.month)] = (d, c)
    chaves = sorted(fim_mes.keys())
    rets = {}
    for i in range(1, len(chaves)):
        _, c_prev = fim_mes[chaves[i-1]]
        _, c_cur  = fim_mes[chaves[i]]
        rets[chaves[i]] = (c_cur / c_prev) - 1
    return rets


def retornos_mensais_ativo(nome: str, filepath: Path = QUANTUM_FILE) -> dict:
    series = _load_series(filepath)
    return _retornos_mensais(series.get(nome, []))


def calcular_portfolio(aloc_fundo_pct: dict, filepath: Path = QUANTUM_FILE) -> dict:
    """
    aloc_fundo_pct: {nome_quantum: pct_do_total (0-1)}
    Retorna série mensal do portfólio ponderado + estatísticas.
    Fundos sem histórico em um mês são renormalizados.
    """
    series = _load_series(filepath)

    rets_por_fundo = {}
    for nome, w in aloc_fundo_pct.items():
        if w <= 0:
            continue
        rets_por_fundo[nome] = _retornos_mensais(series.get(nome, []))

    cdi_rets = _retornos_mensais(series.get("CDI", []))

    # Meses onde há pelo menos um fundo com dado
    todos_meses = set()
    for rets in rets_por_fundo.values():
        todos_meses.update(rets.keys())
    todos_meses = sorted(todos_meses)

    port_rets = {}
    for m in todos_meses:
        soma_w, soma_r = 0.0, 0.0
        for nome, w in aloc_fundo_pct.items():
            if w <= 0: continue
            r = rets_por_fundo.get(nome, {}).get(m)
            if r is not None:
                soma_w += w
                soma_r += w * r
        if soma_w > 0.3:  # exigir cobertura mínima de 30% p/ mês válido
            port_rets[m] = soma_r / soma_w

    # Estatísticas
    meses_ord = sorted(port_rets.keys())
    if not meses_ord:
        return {}

    acum = 1.0
    for m in meses_ord:
        acum *= (1 + port_rets[m])

    n_meses = len(meses_ord)
    ret_total = acum - 1
    ret_anual = (acum ** (12 / n_meses)) - 1 if n_meses >= 6 else None

    import statistics
    rets_lst = [port_rets[m] for m in meses_ord]
    vol_mensal = statistics.pstdev(rets_lst) if len(rets_lst) > 2 else None
    vol_anual  = vol_mensal * math.sqrt(12) if vol_mensal else None

    cdi_acum = 1.0
    cdi_meses = 0
    for m in meses_ord:
        if m in cdi_rets:
            cdi_acum *= (1 + cdi_rets[m])
            cdi_meses += 1
    cdi_total = cdi_acum - 1 if cdi_meses else None
    cdi_anual = (cdi_acum ** (12 / cdi_meses)) - 1 if cdi_meses >= 6 else None

    sharpe = None
    if ret_anual is not None and cdi_anual is not None and vol_anual:
        sharpe = (ret_anual - cdi_anual) / vol_anual

    pos = sum(1 for r in rets_lst if r > 0)
    neg = sum(1 for r in rets_lst if r < 0)
    acima_cdi  = sum(1 for m in meses_ord
                     if m in cdi_rets and port_rets[m] > cdi_rets[m])
    abaixo_cdi = sum(1 for m in meses_ord
                     if m in cdi_rets and port_rets[m] <= cdi_rets[m])

    return {
        "retornos_mensais": port_rets,       # {(ano,mes): ret}
        "cdi_mensais":      cdi_rets,
        "ret_total":        ret_total,
        "ret_anualizado":   ret_anual,
        "vol_anualizada":   vol_anual,
        "cdi_total":        cdi_total,
        "pct_cdi":          (ret_total / cdi_total) if cdi_total else None,
        "sharpe":           sharpe,
        "meses_positivos":  pos,
        "meses_negativos":  neg,
        "maior_retorno":    max(rets_lst),
        "menor_retorno":    min(rets_lst),
        "acima_cdi":        acima_cdi,
        "abaixo_cdi":       abaixo_cdi,
        "n_meses":          n_meses,
    }
